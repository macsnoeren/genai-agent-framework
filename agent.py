#!/usr/bin/env python

import sys
from pathlib import Path
from typing import Dict, Any, Optional, List, Union
from datetime import datetime
import logging
import requests
import json
import shutil
import argparse

# Voorkom UnicodeEncodeErrors door UTF-8 te forceren voor de console output
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding='utf-8')

# Voeg de 'lib' directory toe aan sys.path, als deze niet al is toegevoegd
if str(Path(__file__).parent / "lib") not in sys.path:
    sys.path.append(str(Path(__file__).parent / "lib"))

from lib.docdialog_client import DocumentDialogueClient
from lib.ollama_client import OllamaClient
from lib.ai_agent import AIAgent

# Configureer logging
logging.basicConfig(
    level=logging.INFO, 
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("data/agent_log.log", encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# Definieer een aparte logger voor de agent's interne stappen
agent_logger = logging.getLogger('RWEAIAgent')
agent_logger.setLevel(logging.INFO) # Kan op DEBUG gezet worden voor meer details

def load_config():
    config_path = Path(__file__).parent / "config.json"
    if config_path.exists():
        with open(config_path, 'r') as f:
            return json.load(f)
    return {}

def extract_file_content(file_path: Path) -> str:
    """
    Extraheert tekst uit .txt, .docx en .xlsx bestanden voor gebruik in de prompt.
    """
    ext = file_path.suffix.lower()
    try:
        if ext == '.txt':
            return file_path.read_text(encoding='utf-8')
        elif ext == '.docx':
            import docx  # pip install python-docx
            doc = docx.Document(file_path)
            return "\n".join([para.text for para in doc.paragraphs])
        elif ext == '.xlsx':
            import openpyxl  # pip install openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            content = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                content.append(f"--- Tabblad: {sheet_name} ---")
                for row in ws.iter_rows(values_only=True):
                    # Converteer rij naar CSV-stijl string
                    line = ",".join([str(cell) if cell is not None else "" for cell in row])
                    content.append(line)
            return "\n".join(content)
        else:
            logger.warning(f"Bestandstype {ext} wordt niet ondersteund voor tekstextractie.")
            return ""
    except Exception as e:
        logger.error(f"Fout bij extraheren van {file_path.name}: {e}")
        return ""

def load_agent_config(config_filename: str):
    config_path = Path(config_filename)
    # Als het bestand niet direct gevonden wordt, zoek dan in de script directory
    if not config_path.exists():
        config_path = Path(__file__).parent / config_filename

    if config_path.exists():
        with open(config_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    raise FileNotFoundError(f"Configuratiebestand '{config_filename}' niet gevonden.")

def generate_report(data: Dict[str, Any], template_path: Path, output_path: Path):
    """
    Genereert een Word-document op basis van een template en JSON data.
    """
    try:
        from docxtpl import DocxTemplate
        if not template_path.exists():
            logger.warning(f"Template niet gevonden op {template_path}, overslaan rapportage.")
            return

        doc = DocxTemplate(template_path)
        doc.render(data)
        doc.save(output_path)
        print(f"Rapport gegenereerd: {output_path.name}")
    except ImportError:
        logger.error("docxtpl niet geïnstalleerd. Run: pip install docxtpl")
    except Exception as e:
        logger.error(f"Fout bij genereren rapport: {e}")

def run_agent_batch(config_filename: str, access_token: str, global_config: Dict[str, Any]):
    """
    Voert de batchverwerking uit voor een specifieke agent configuratie.
    """
    agent_config = load_agent_config(config_filename)

    # Paden instellen vanuit configuratie
    input_dir = Path(agent_config.get("input_directory", "data/input"))
    output_dir = Path(agent_config.get("output_directory", "data/output"))
    done_dir = Path(agent_config.get("done_directory", "data/done"))
    report_dir = Path(agent_config.get("report_directory", "data/reports"))

    # Zorg dat mappen bestaan
    for d in [input_dir, output_dir, done_dir, report_dir]:
        try:
            d.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            logger.info(f"Informatie: Kon map {d} niet aanmaken of benaderen: {e}")

    files = [f for f in input_dir.rglob('*') if f.is_file()]
    if not files:
        print(f"Geen bestanden gevonden in {input_dir}")
        return

    # --- LLM Initialisatie (pas na vinden van bestanden) ---
    provider_type = agent_config.get("provider", "ollama").lower()
    if provider_type == "docdialog":
        if access_token == "VERVANG_DOOR_JE_ECHTE_TOKEN":
            logger.error(f"DocumentDialogue geselecteerd voor {config_filename} maar geen geldige ACCESS_TOKEN gevonden.")
            return
        client = DocumentDialogueClient(access_token)
    else:
        ollama_url = global_config.get("OLLAMA_BASE_URL", "http://localhost:11434")
        client = OllamaClient(base_url=ollama_url)
    
    agent = AIAgent(client)
    print(f"Agent gestart met provider: {type(client).__name__}")

    print("Beschikbare modellen ophalen...")
    models = agent.list_models()
    
    selected_model = agent_config.get("model")
    if selected_model:
        print(f"Model geselecteerd uit configuratie: {selected_model}")
    elif models:
        selected_model = models[0].get("id")
        print(f"Geen model opgegeven in agent.json. Automatisch eerste model geselecteerd: {selected_model}")
    else:
        print("Geen modellen gevonden op de client. We gaan door zonder specifiek model.")
        selected_model = None

    print(f"\nStart verwerking van {len(files)} bestanden...")

    for file_path in files:
        rel_path = file_path.relative_to(input_dir)
        print(f"\n--- Bezig met: {rel_path} ---")

        # Start een schone chat per bestand
        print("Nieuwe chat aanmaken...")
        chat_id = agent.create_chat(model_id=selected_model)
        print(f"Chat aangemaakt met ID: {chat_id}")
        logger.info(f"Actieve chat_id voor agent: {chat_id}")

        try:
            # 4. Bestand uploaden of inhoud extraheren als fallback
            file_content = ""
            try:
                agent.upload_file(str(file_path))
            except (NotImplementedError, Exception) as e:
                logger.info(f"Upload niet mogelijk voor {file_path.name} ({e}). Inhoud wordt handmatig geëxtraheerd.")
                file_content = extract_file_content(file_path)

            # Stel de taak samen voor de agent
            task_prompt = (
                f"Instructie: {agent_config['instructions']}\n"
                f"Bestand om te verwerken: {file_path.name}\n"
                f"Verwachte output formaat: {agent_config['output_description']}\n"
            )

            if file_content:
                task_prompt += f"\nInhoud van het bestand:\n---\n{file_content}\n---\n"
            
            task_prompt += "\nGeef alleen de rauwe JSON terug in je uiteindelijke antwoord."

            # Gebruik de run_agent methode voor multi-step reasoning
            max_iterations = agent_config.get("max_iterations", 3)
            response = agent.run_agent(task_prompt, max_iterations=max_iterations)
            content = response.get("content", "")

            # 5. Output opslaan als JSON (behoud structuur)
            output_file = output_dir / rel_path.with_suffix('.json')
            output_file.parent.mkdir(parents=True, exist_ok=True)
            output_file.write_text(content, encoding='utf-8')

            # 5b. Optioneel rapport genereren in huisstijl
            template_path = agent_config.get("template_path")
            if template_path:
                try:
                    # Parse de content om er zeker van te zijn dat het valide JSON is
                    json_data = json.loads(content)
                    report_path = report_dir / rel_path.with_suffix('.docx')
                    report_path.parent.mkdir(parents=True, exist_ok=True)
                    generate_report(json_data, Path(template_path), report_path)
                except json.JSONDecodeError:
                    logger.warning(f"Kon content voor {file_path.name} niet naar JSON parsen voor rapportage.")
            
            # 5c. Optioneel data toevoegen aan de verzamelbak (collection file)
            collection_path_str = agent_config.get("collection_file_path")
            if collection_path_str:
                try:
                    json_data = json.loads(content)
                    db_entry = json_data.get("database_file")
                    if db_entry:
                        collection_path = Path(collection_path_str)
                        collection_path.parent.mkdir(parents=True, exist_ok=True)
                        with open(collection_path, "a", encoding="utf-8") as f:
                            f.write(json.dumps(db_entry) + "\n")
                        print(f"Data toegevoegd aan verzamelbak: {collection_path.name}")
                except Exception as e:
                    logger.error(f"Fout bij bijwerken verzamelbak voor {file_path.name}: {e}")

            # 6. Verplaatsen naar de 'done' directory
            if done_dir.exists():
                try:
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
                    target_done_filename = f"{timestamp}_{file_path.name}"
                    target_done_path = done_dir / rel_path.parent / target_done_filename

                    target_done_path.parent.mkdir(parents=True, exist_ok=True)
                    shutil.move(str(file_path), str(target_done_path))
                    print(f"Succesvol verwerkt: {output_file.name}")
                except Exception as e:
                    logger.info(f"Informatie: Bestand kon niet worden verplaatst naar done: {e}")
                    print(f"Succesvol verwerkt (niet verplaatst naar done): {output_file.name}")
            else:
                logger.info(f"Informatie: Done directory '{done_dir}' bestaat niet. Bestand is niet verplaatst.")
                print(f"Succesvol verwerkt (done map ontbreekt): {output_file.name}")

        except Exception as e:
            logger.error(f"Fout bij verwerken van {file_path.name}: {e}")
            print(f"Fout bij {file_path.name}: {e}")

        # Sluit de chat af na verwerking van het bestand
        agent.delete_current_chat()
        print(f"Chat {chat_id} gesloten.")

    print("\nBatch-verwerking voltooid.")

def main():
    parser = argparse.ArgumentParser(description="Draai de AI Agent batch verwerking met een specifieke configuratie.")
    parser.add_argument(
        "config", 
        nargs="?", 
        help="Pad naar het agent JSON configuratiebestand. Indien niet opgegeven worden alle JSON bestanden in 'agents/' uitgevoerd."
    )
    args = parser.parse_args()

    global_config = load_config()
    access_token = global_config.get("ACCESS_TOKEN", "VERVANG_DOOR_JE_ECHTE_TOKEN")

    if args.config:
        config_files = [args.config]
    else:
        # Scan de agents directory naar JSON bestanden
        agents_dir = Path(__file__).parent / "agents"
        config_files = [str(f) for f in agents_dir.glob("*.json")]
        
        if not config_files:
            logger.error("Geen agent configuratiebestanden gevonden in de 'agents/' directory.")
            return

    for config_file in config_files:
        print(f"\n" + "="*50)
        print(f"START VERWERKING AGENT: {config_file}")
        print("="*50 + "\n")
        try:
            run_agent_batch(config_file, access_token, global_config)
        except Exception as e:
            logger.error(f"Fout bij verwerken van agent {config_file}: {e}")

if __name__ == "__main__":
    main()
